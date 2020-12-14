import matplotlib.pyplot as plt
import numpy as np
import seaborn as sns
import pandas as pd
import sys
import openpyxl
import os
import math

os.chdir('D:\\Desktop')

def t1(x1,x2):
	return x2-0.1 #0 degree polynomial
def t2(x1,x2):
        return x2-1.5*x1+0.05 #first degree polynomial
def t3(x1,x2):
        return x2-18*(x1**2)*+1.3*x1+1.4 #second degree polynomial
def t4(x1,x2):
        return x2-172.8*(x1**3)+40*(x1**2)+x1-0.25 #third degree polynomial
def t5(x1,x2):
        if x1<-0.2:
                return t2(x1,x2)
        elif x1<0.2:
                return t4(x1,x2)
        else:
                return -(t3(x1,x2))#different functions in different ranges
def t6(x1,x2):
        if (x1-0.2)**2+(x2+0.3)**2<0.1**2:
                return 1
        elif (x1+0.4)**2-(x2-0.1)**3 < 0.1:
                return -1
        else:
                return t4(x1,x2) #logical combinations of different function
def t7(x1,x2):
        if 0.01*(x2**2) > 100*((x1+2.5)**2)*math.sin(100*((x1+2.5)**2)+0.01*x1**2):
                return 1
        else:
                return -1 #complex trig function

targetFunctions=[t1,t2,t3,t4,t5,t6,t7]
maxError=[]

for t in range(0,7):
    err1=0
    err2=0
    INDEX1=-0.5
    while INDEX1<=0.5:
        INDEX2=-0.5
        while INDEX2<=0.5:
            if(targetFunctions[t])(INDEX1, INDEX2) >= 0:
                err1 += 1
            else:
                err2 += 1
            INDEX2+=0.02
        INDEX1+=0.02
    me=err1/2601
    if err2>err1:
        me=err2/2601
    maxError.extend([me])

def noiseLevelDiagrams(num):
    wb = openpyxl.load_workbook("excel" + str(num) + ".xlsx")
    sheet = wb.active

    error=[]
    noiseLevels=[]
    funcError=[ [], [], [], [], [], [] ]
    optC=[ [], [], [], [], [], [] ]

    for i in range(1, 16500, 1500):
        error.extend([maxError[0]])
        noiseLevels.extend([float(sheet.cell(row=i, column=1).value)])
        sum1=[0,0,0,0,0,0,0]
        sum2=[0,0,0,0,0,0,0]
        rem=[0,0,0,0,0,0,0]
        for j in range(0,1500,6):
            for k in range(0,6):
                if((float(sheet.cell(row=i+j+k, column=7).value) != 1)  and (sheet.cell(row=i+j+k, column=6).value != "undefined")):
                    sum1[k]+=float(sheet.cell(row=i+j+k, column=7).value)
                    sum2[k]+=float(sheet.cell(row=i+j+k, column=6).value)
                else:
                    rem[k]+=1
        for j in range(0,6):
            '''if(rem[j] == 250):
                func.Error[j].extend([2*funcError[j-1]-funcError[j-2]])
                optC.extend([2*optC[j-1]-optC[j-2]])
                continue'''
            sum1[j] /= (250-rem[j])
            sum2[j] /= (250-rem[j])
            funcError[j].extend([sum1[j]])
            optC[j].extend([sum2[j]])
            
    df=pd.DataFrame({'Noise levels':noiseLevels, 'Linear': funcError[0],
                     'Polynomial2': funcError[1], 'Polynomial3': funcError[2],
                     'Polynomial5': funcError[3], 'Polynomial10': funcError[4],
                     'Polynomial100': funcError[5], 'Linear ': optC[0],
                     'Polynomial2 ': optC[1], 'Polynomial3 ': optC[2],
                     'Polynomial5 ': optC[3], 'Polynomial10 ': optC[4],
                     'Polynomial100 ': optC[5]})

    plt.figure(0)
    plt.plot('Noise levels', 'Linear', color='blue', linewidth=1, data=df)
    plt.plot('Noise levels', 'Polynomial2', color='green', linewidth=1.5, data=df)
    plt.plot('Noise levels', 'Polynomial3', color='red', linewidth=2, data=df)
    plt.plot('Noise levels', 'Polynomial5', color='cyan', linewidth=2.5, data=df)
    plt.plot('Noise levels', 'Polynomial10', color='magenta', linewidth=3, data=df)
    plt.plot('Noise levels', 'Polynomial100', color='yellow', linewidth=3.5, data=df)
    """plt.plot('Noise levels', 'Max Error', color='black', linewidth=1, linestyle='dashed', data=df)"""
    plt.legend()
    plt.title("t" + str(num+1) + " with max error " + str(maxError[num]))
    plt.xlabel("Noise")
    plt.ylabel("Average error")
    plt.show()

    plt.figure(1)
    plt.plot('Noise levels', 'Linear ', color='blue', linewidth=1, data=df)
    plt.plot('Noise levels', 'Polynomial2 ', color='green', linewidth=1.5, data=df)
    plt.plot('Noise levels', 'Polynomial3 ', color='red', linewidth=2, data=df)
    plt.plot('Noise levels', 'Polynomial5 ', color='cyan', linewidth=2.5, data=df)
    plt.plot('Noise levels', 'Polynomial10 ', color='magenta', linewidth=3, data=df)
    plt.plot('Noise levels', 'Polynomial100 ', color='yellow', linewidth=3.5, data=df)
    plt.legend()
    plt.title("t"+str(num+1))
    plt.xlabel("Noise")
    plt.ylabel("Optimal C")
    plt.show()

noiseLevelDiagrams(0)
noiseLevelDiagrams(1)
noiseLevelDiagrams(2)
noiseLevelDiagrams(3)
noiseLevelDiagrams(4)
noiseLevelDiagrams(5)
noiseLevelDiagrams(6)

def numberOfExampleDiagrams(num):
        wb=openpyxl.load_workbook("excel" + str(num) + ".xlsx")
        sheet=wb.active

        error=[]
        numberOfExamples=[]
        funcError=[ [], [], [], [], [], [] ]
        optC=[ [], [], [], [], [], [] ]

        for f in range(1,1500,30):
                numberOfExamples.extend([float(sheet.cell(row=f, column=2).value)])
                sum1=[0,0,0,0,0,0,0]
                sum2=[0,0,0,0,0,0,0]
                rem=[0,0,0,0,0,0,0]
                for i in range(0, 16500, 1500):
                        for j in range(0,30,6):
                                for k in range(0,6):
                                        if((float(sheet.cell(row=f+i+j+k, column=7).value != 1)) and (sheet.cell(row=f+i+j+k, column=6).value != "undefined")):
                                                sum1[k]+=float(sheet.cell(row=f+i+j+k, column=7).value)
                                                sum2[k]+=float(sheet.cell(row=f+i+j+k, column=6).value)
                                        else:
                                                rem[k]+=1
                for i in range(0,6):
                        if rem[i]==55:
                                '''funcError[i].extend([2*funcError[i][len(funcError[i])-1]-funcError[i][len(funcError[i])-2]])
                                optC[i].extend([2*optC[i][len(optC[i])-1]-optC[i][len(optC[i])-2]])'''
                                funcError[i].extend([funcError[i][len(funcError[i])-2]])
                                optC[i].extend([optC[i][len(optC[i])-2]])
                                continue
                        sum1[i] /= (55-rem[i])
                        sum2[i] /= (55-rem[i])
                        funcError[i].extend([sum1[i]])
                        optC[i].extend([sum2[i]])
                        
        df=pd.DataFrame({'Number of Examples':numberOfExamples, 'Linear': funcError[0],
                        'Polynomial2': funcError[1], 'Polynomial3': funcError[2],
                        'Polynomial5': funcError[3], 'Polynomial10': funcError[4],
                        'Polynomial100': funcError[5], 'Linear ': optC[0],
                        'Polynomial2 ': optC[1], 'Polynomial3 ': optC[2],
                        'Polynomial5 ': optC[3], 'Polynomial10 ': optC[4],
                        'Polynomial100 ': optC[5]})

        plt.figure(0)
        plt.plot('Number of Examples', 'Linear', color='blue', linewidth=1, data=df)
        plt.plot('Number of Examples', 'Polynomial2', color='green', linewidth=1.5, data=df)
        plt.plot('Number of Examples', 'Polynomial3', color='red', linewidth=2, data=df)
        plt.plot('Number of Examples', 'Polynomial5', color='cyan', linewidth=2.5, data=df)
        plt.plot('Number of Examples', 'Polynomial10', color='magenta', linewidth=3, data=df)
        plt.plot('Number of Examples', 'Polynomial100', color='yellow', linewidth=3.5, data=df)
        """plt.plot('Noise levels', 'Max Error', color='black', linewidth=1, linestyle='dashed', data=df)"""
        plt.legend()
        plt.title("t" + str(num+1) + " with max error " + str(maxError[num]))
        plt.xlabel("Number of Examples")
        plt.ylabel("Average error")
        plt.show()

        plt.figure(1)
        plt.plot('Number of Examples', 'Linear ', color='blue', linewidth=1, data=df)
        plt.plot('Number of Examples', 'Polynomial2 ', color='green', linewidth=1.5, data=df)
        plt.plot('Number of Examples', 'Polynomial3 ', color='red', linewidth=2, data=df)
        plt.plot('Number of Examples', 'Polynomial5 ', color='cyan', linewidth=2.5, data=df)
        plt.plot('Number of Examples', 'Polynomial10 ', color='magenta', linewidth=3, data=df)
        plt.plot('Number of Examples', 'Polynomial100 ', color='yellow', linewidth=3.5, data=df)
        plt.legend()
        plt.title("t"+str(num+1))
        plt.xlabel("Number of Examples")
        plt.ylabel("Optimal C")
        plt.show()

numberOfExampleDiagrams(0)
numberOfExampleDiagrams(1)
numberOfExampleDiagrams(2)
numberOfExampleDiagrams(3)
numberOfExampleDiagrams(4)
numberOfExampleDiagrams(5)
numberOfExampleDiagrams(6)
