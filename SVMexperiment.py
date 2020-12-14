from random import seed
from random import random
from sklearn import svm
import numpy as np
import sys
import openpyxl
import os
import math
import warnings

os.chdir('D:\\Desktop')
wb=[openpyxl.Workbook()]
acIND=0
for i in range(6):
        wb.append(openpyxl.Workbook())
def t1(x1,x2):
	return x2-0.1 #0 degree polynomial
def t2(x1,x2):
        return x2-1.5*x1+0.05 #first degree polynomial
def t3(x1,x2):
        return x2-18*(x1**2)*+1.3*x1+1.4 #second degree polynomial
def t4(x1,x2):
        return x2-172.8*(x1**3)+40*(x1**2)+x1-0.25 #third degree polynomial
def t5(x1,x2):
        if x1<-0.2:
                return t2(x1,x2)
        elif x1<0.2:
                return t4(x1,x2)
        else:
                return -(t3(x1,x2))#different functions in different ranges
def t6(x1,x2):
        if (x1-0.2)**2+(x2+0.3)**2<0.1**2:
                return 1
        elif (x1+0.4)**2-(x2-0.1)**3 < 0.1:
                return -1
        else:
                return t4(x1,x2) #logical combinations of different function
def t7(x1,x2):
        if 0.01*(x2**2) > 100*((x1+2.5)**2)*math.sin(100*((x1+2.5)**2)+0.01*x1**2):
                return 1
        else:
                return -1 #complex trig function
def linear(x1,x2):
	return np.dot(x1,x2.T)
def polynomial2(x1,x2):
	return (linear(x1,x2)+1)**2
def polynomial3(x1,x2):
	return (linear(x1,x2)+1)**3
def polynomial5(x1,x2):
	return (linear(x1,x2)+1)**5
def polynomial10(x1,x2):
	return (linear(x1,x2)+1)**10
def polynomial100(x1,x2):
	return (linear(x1,x2)+1)**100
targetFunctions = [t1, t2, t3, t4, t5, t6, t7]
functionError=[]
for i in range(7):
        functionError.append([0,0])
for t in range(0,7):
        INDEX1=-0.5
        while INDEX1<=0.5:
                INDEX2=-0.5
                while INDEX2<=0.5:
                        if (targetFunctions[t])(INDEX1, INDEX2) >= 0:
                                functionError[t][0] += 1
                        else:
                                functionError[t][1] +=1
                        INDEX2+=0.02
                INDEX1+=0.02
        functionError[t][0]/=2601
        functionError[t][1]/=2601
kernels=[linear, polynomial2, polynomial3, polynomial5, polynomial10,polynomial100]
possC=[]
for i in functionError:
        print(str(i[0]) + " " + str(i[1]))
for i in range(-4,3):
        possC.append(10**i)
        possC.append(5 * (10**i))
seed(1)
genError=0
for t in range(0,len(targetFunctions)):
        noiseRow=1
        NRow=1
        exRow=1
        kernelRow=1
        ac=wb[acIND]
        acIND+=1
        sh=ac.active
        NOISE=0
        while NOISE<=1:
                sh.cell(row=noiseRow, column=1).value=NOISE
                noiseRow+=1500
                N=10
                while N<=1000:
                        sh.cell(row=NRow, column=2).value=N
                        NRow+=30
                        for temp in range(5):
                                y=[]
                                ex=[]
                                for tempp in range(N):
                                        x1=random()-0.5
                                        x2=random()-0.5
                                        if targetFunctions[t](x1,x2) >= 0:
                                                y.append(1)
                                        else:
                                                y.append(0)
                                        r1=(random()*NOISE/2)+(NOISE/2)
                                        if random() < 0.5:
                                                r1 = -r1
                                        r2=(random()*NOISE/2)+(NOISE/2)
                                        if random() < 0.5:
                                                r2-=r2
                                        ex.append([x1*(1+r1), x2*(1+r2)])
                                sh.cell(row=exRow, column=3).value=str(ex)
                                sh.cell(row=exRow, column=4).value=str(y)
                                exRow+=6
                                uniqueClass=True
                                for tempp in y:
                                        if tempp != y[0]:
                                                uniqueClass=False
                                                break
                                if uniqueClass:
                                        if y[0]==1:
                                                genError=functionError[t][1]
                                        else:
                                                genError=functionError[t][0]
                                        for ker in kernels:
                                                sh.cell(row=kernelRow, column=5).value=str(ker)
                                                sh.cell(row=kernelRow, column=6).value="undefined"
                                                sh.cell(row=kernelRow, column=7).value=genError
                                                kernelRow+=1
                                        continue
                                for ker in kernels:
                                        sh.cell(row=kernelRow, column=5).value=str(ker)
                                        minError=-1
                                        perfectC=-1
                                        for PC in possC:
                                                error=0
                                                for COUNT in range(10):
                                                        val=ex[COUNT*N//10 : (COUNT+1)*N//10]
                                                        tr = ex[0:COUNT*N//10]
                                                        tr += ex[(COUNT+1)*N//10 : N]
                                                        valY=y[COUNT*N//10 : (COUNT+1)*N//10]
                                                        trY = y[0:COUNT*N//10]
                                                        trY += y[(COUNT+1)*N//10 : N]
                                                        unCl=True
                                                        for temporary in trY:
                                                                if temporary!=trY[0]:
                                                                        unCl=False
                                                                        break
                                                        if unCl:
                                                                for dis in valY:
                                                                        if dis!=y[0]:
                                                                                error+=1
                                                                continue
                                                        with warnings.catch_warnings(record=True) as w:
                                                                warnings.simplefilter("error")
                                                                try:
                                                                        clf = svm.SVC(kernel=ker, C=PC, max_iter=1000000)
                                                                        clf.fit(tr, trY)
                                                                except:
                                                                        error+=N//10
                                                                        continue
                                                        for cur in range(N//10):
                                                                if clf.predict([val[cur]]) != valY[cur]:
                                                                        error+=1
                                                if ((error < minError) or (minError<0)):
                                                        minError=error
                                                        perfectC=PC
                                        with warnings.catch_warnings(record=True) as w:
                                                warnings.simplefilter("error")
                                                sh.cell(row=kernelRow, column=6).value = perfectC
                                                try:
                                                        clf=svm.SVC(kernel=ker, C=perfectC, max_iter=1000000)
                                                        clf.fit(ex,y)
                                                except:
                                                        sh.cell(row=kernelRow, column=7).value=1
                                                        kernelRow+=1
                                                        continue
                                        misclassified=0
                                        INDEX1=-0.5
                                        while INDEX1<=0.5:
                                                INDEX2=-0.5
                                                while INDEX2<=0.5:
                                                        expectation=0
                                                        if targetFunctions[t](INDEX1,INDEX2) >= 0:
                                                                expectation=1
                                                        if clf.predict([[INDEX1, INDEX2]]) != expectation:
                                                                misclassified += 1
                                                        INDEX2 += 0.02
                                                INDEX1 += 0.02
                                        sh.cell(row=kernelRow,column=7).value=misclassified/2601
                                        kernelRow+=1
                        N+=20
                NOISE+=0.1
        ac.save("excel" + str(acIND-1) + ".xlsx")
